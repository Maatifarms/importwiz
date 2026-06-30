from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from buyer_data import BUYERS

wb = Workbook()

HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HOT_FILL = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
WARM_FILL = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')
GREEN_FILL = PatternFill(start_color='6BCB77', end_color='6BCB77', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

MAIN_HEADERS = [
    "Rank", "Company Name", "Website", "Country", "Address",
    "Decision Maker", "Designation", "Email", "Phone",
    "LinkedIn", "Instagram", "Facebook",
    "Revenue", "Employees", "Years in Business",
    "Product Categories", "Buys Plush?", "From Country",
    "Buying Frequency", "Private Label?", "OEM?",
    "Seasonal Products?", "Target Customer", "Retail Price Range",
    "Import History", "Supplier Countries", "Current Brands",
    "Buying Signal", "Intent Score",
    "Why Buy Now", "Pain Point We Solve",
    "Best Product to Pitch", "Expected Order Size", "Urgency Level"
]

INTEL_HEADERS = [
    "Rank", "Company Name", "Country", "Intent Score",
    "Email Opening", "LinkedIn Message", "WhatsApp Message", "Follow-Up Strategy"
]

PITCH_HEADERS = [
    "Rank", "Company Name", "Country", "Intent Score",
    "Why They Buy Now", "Pain Point", "Best Product",
    "Expected Order", "Urgency", "Pitch Approach"
]

def style_header(ws, headers):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 35
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

def write_buyer_row(ws, row_num, rank, b):
    data = [
        rank, b[0], b[1], b[2], b[3], b[4], b[5], b[6], b[7],
        b[8], b[9], b[10], b[11], b[12], b[13], b[14], b[15],
        b[16], b[17], b[18], b[19], b[20], b[21], b[22], b[23],
        b[24], b[25], b[26], b[27], b[28], b[29], b[30], b[31], b[32]
    ]
    for col_idx, val in enumerate(data, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        cell.border = THIN_BORDER
    score = b[27]
    score_cell = ws.cell(row=row_num, column=29)
    if score >= 93:
        score_cell.fill = HOT_FILL
    elif score >= 90:
        score_cell.fill = WARM_FILL
    elif score >= 85:
        score_cell.fill = GREEN_FILL

def generate_email(b):
    company = b[0]
    dm = b[4] if b[4] != "Not Verified" else "Procurement Team"
    product = b[30]
    pain = b[29]
    return (
        f"Subject: Premium Plush Toys - EN71/ASTM Certified Manufacturer from India | Namaste Overseas\n\n"
        f"Dear {dm},\n\n"
        f"I noticed that {company} has been actively sourcing plush toys and I believe Namaste Overseas "
        f"can be a valuable manufacturing partner for your team.\n\n"
        f"We specialize in {product} with full EN71, ASTM & CE certification, child-safe materials, "
        f"and MOQ starting from just 100 pieces.\n\n"
        f"We understand that {pain.lower()} - and this is exactly where we can help. "
        f"Our India-based manufacturing offers 15-30% cost advantages over China with comparable quality, "
        f"zero tariff risks, and complete export-ready packaging.\n\n"
        f"Would you be open to a quick 15-minute call to discuss how we can support your upcoming collections?\n\n"
        f"Best regards,\nNamaste Overseas\nIndia"
    )

def generate_linkedin(b):
    company = b[0]
    dm = b[4] if b[4] != "Not Verified" else "Hi"
    product = b[30]
    return (
        f"Hi {dm},\n\n"
        f"I came across {company} and was impressed by your product range. "
        f"At Namaste Overseas (India), we manufacture premium {product} - "
        f"EN71/ASTM/CE certified, MOQ from 100 pcs, with private label & OEM capability.\n\n"
        f"India offers significant cost advantages and supply chain diversification vs China. "
        f"Would love to explore a partnership. Open to connecting?"
    )

def generate_whatsapp(b):
    company = b[0]
    product = b[30]
    return (
        f"Hi! This is Namaste Overseas from India. We're a certified manufacturer of {product}. "
        f"We noticed {company} sources plush toys and wanted to introduce ourselves. "
        f"EN71/ASTM/CE certified, MOQ 100 pcs, private label & OEM available. "
        f"Can I send you our catalog?"
    )

def generate_followup(b):
    score = b[27]
    if score >= 93:
        return "Day 1: Email + LinkedIn connect | Day 3: Follow-up email with catalog PDF | Day 5: WhatsApp with sample photos | Day 7: Call attempt | Day 14: Send physical sample pack | Day 21: Schedule video call for factory tour"
    elif score >= 90:
        return "Day 1: Email + LinkedIn connect | Day 4: Follow-up email with catalog | Day 7: WhatsApp introduction | Day 14: Call or video proposal | Day 21: Send sample pack"
    elif score >= 85:
        return "Day 1: Email introduction | Day 5: LinkedIn connect | Day 10: Follow-up email | Day 20: WhatsApp with new collection images | Day 30: Re-engage with seasonal pitch"
    else:
        return "Day 1: Email introduction | Day 7: LinkedIn connect | Day 14: Follow-up email | Day 30: Seasonal re-engagement | Quarterly: Newsletter updates"

def auto_width(ws, max_width=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 12)

sorted_buyers = sorted(BUYERS, key=lambda x: x[27], reverse=True)

# Sheet 1: TOP 100 BUYERS
ws1 = wb.active
ws1.title = "TOP 100 BUYERS"
style_header(ws1, MAIN_HEADERS)
for i, b in enumerate(sorted_buyers[:100]):
    write_buyer_row(ws1, i + 2, i + 1, b)
auto_width(ws1)

# Sheet 2: TOP 25 HOT BUYERS
ws2 = wb.create_sheet("TOP 25 HOT BUYERS")
style_header(ws2, MAIN_HEADERS)
for i, b in enumerate(sorted_buyers[:25]):
    write_buyer_row(ws2, i + 2, i + 1, b)
    for col in range(1, len(MAIN_HEADERS) + 1):
        ws2.cell(row=i + 2, column=col).fill = LIGHT_BLUE
    ws2.cell(row=i + 2, column=29).fill = HOT_FILL
auto_width(ws2)

# Sheet 3: TOP 10 IMMEDIATE BUYERS
ws3 = wb.create_sheet("TOP 10 IMMEDIATE")
style_header(ws3, MAIN_HEADERS)
for i, b in enumerate(sorted_buyers[:10]):
    write_buyer_row(ws3, i + 2, i + 1, b)
    for col in range(1, len(MAIN_HEADERS) + 1):
        ws3.cell(row=i + 2, column=col).fill = LIGHT_BLUE
    ws3.cell(row=i + 2, column=29).fill = HOT_FILL
auto_width(ws3)

# Sheet 4: TOP 5 ORDER THIS MONTH
ws4 = wb.create_sheet("TOP 5 ORDER THIS MONTH")
style_header(ws4, MAIN_HEADERS)
for i, b in enumerate(sorted_buyers[:5]):
    write_buyer_row(ws4, i + 2, i + 1, b)
    for col in range(1, len(MAIN_HEADERS) + 1):
        ws4.cell(row=i + 2, column=col).fill = LIGHT_BLUE
    ws4.cell(row=i + 2, column=29).fill = HOT_FILL
auto_width(ws4)

# Sheet 5: BUYER INTELLIGENCE (with outreach templates)
ws5 = wb.create_sheet("BUYER INTELLIGENCE")
style_header(ws5, INTEL_HEADERS)
for i, b in enumerate(sorted_buyers[:100]):
    row = i + 2
    ws5.cell(row=row, column=1, value=i + 1).border = THIN_BORDER
    ws5.cell(row=row, column=2, value=b[0]).border = THIN_BORDER
    ws5.cell(row=row, column=3, value=b[2]).border = THIN_BORDER
    ws5.cell(row=row, column=4, value=b[27]).border = THIN_BORDER
    ws5.cell(row=row, column=5, value=generate_email(b)).border = THIN_BORDER
    ws5.cell(row=row, column=6, value=generate_linkedin(b)).border = THIN_BORDER
    ws5.cell(row=row, column=7, value=generate_whatsapp(b)).border = THIN_BORDER
    ws5.cell(row=row, column=8, value=generate_followup(b)).border = THIN_BORDER
    for col in range(1, 9):
        ws5.cell(row=row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
    score_cell = ws5.cell(row=row, column=4)
    if b[27] >= 93:
        score_cell.fill = HOT_FILL
    elif b[27] >= 90:
        score_cell.fill = WARM_FILL
auto_width(ws5, max_width=60)
ws5.column_dimensions['E'].width = 80
ws5.column_dimensions['F'].width = 60
ws5.column_dimensions['G'].width = 50
ws5.column_dimensions['H'].width = 60

# Sheet 6: PITCH STRATEGY
ws6 = wb.create_sheet("PITCH STRATEGY")
style_header(ws6, PITCH_HEADERS)
for i, b in enumerate(sorted_buyers[:25]):
    row = i + 2
    pitch = (
        f"LEAD WITH: {b[30]}\n"
        f"CERTIFICATIONS: EN71, ASTM, CE - all compliant\n"
        f"MOQ ADVANTAGE: Starting from 100 pcs (vs typical 1000+ from China)\n"
        f"COST ADVANTAGE: 15-30% lower than China pricing\n"
        f"SUPPLY CHAIN: India = no US/EU tariff risks on Chinese goods\n"
        f"CUSTOMIZATION: Full OEM & Private Label capability\n"
        f"SAMPLES: Ready to send within 7-10 days"
    )
    ws6.cell(row=row, column=1, value=i + 1).border = THIN_BORDER
    ws6.cell(row=row, column=2, value=b[0]).border = THIN_BORDER
    ws6.cell(row=row, column=3, value=b[2]).border = THIN_BORDER
    ws6.cell(row=row, column=4, value=b[27]).border = THIN_BORDER
    ws6.cell(row=row, column=5, value=b[28]).border = THIN_BORDER
    ws6.cell(row=row, column=6, value=b[29]).border = THIN_BORDER
    ws6.cell(row=row, column=7, value=b[30]).border = THIN_BORDER
    ws6.cell(row=row, column=8, value=b[31]).border = THIN_BORDER
    ws6.cell(row=row, column=9, value=b[32]).border = THIN_BORDER
    ws6.cell(row=row, column=10, value=pitch).border = THIN_BORDER
    for col in range(1, 11):
        ws6.cell(row=row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
    score_cell = ws6.cell(row=row, column=4)
    if b[27] >= 93:
        score_cell.fill = HOT_FILL
    elif b[27] >= 90:
        score_cell.fill = WARM_FILL
auto_width(ws6, max_width=50)
ws6.column_dimensions['J'].width = 60

output_path = r"C:\Users\ojhar\importwiz\Namaste_Overseas_HOT_Buyers_Plush_Toys_2026.xlsx"
wb.save(output_path)
print(f"Excel file saved: {output_path}")
print(f"Total buyers: {len(sorted_buyers)}")
print(f"Top 5 by score:")
for b in sorted_buyers[:5]:
    print(f"  {b[27]} - {b[0]} ({b[2]})")
